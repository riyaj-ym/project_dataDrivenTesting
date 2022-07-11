import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheetName):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetName]
    return sheet.max_row


def getColumnCount(file, sheetName):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rowNum, columnNum):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetName]
    return sheet.cell(rowNum, columnNum).value


def writeData(file, sheetName, rowNum, columnNum, data):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetName]
    sheet.cell(rowNum, columnNum).value = data
    workBook.save(file)


def fillGreenColor(file, sheetName, rowNum, columnNum):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetName]
    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(rowNum, columnNum).fill = greenFill
    workBook.save(file)


def fillRedColor(file, sheetName, rowNum, columnNum):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetName]
    redFill = PatternFill(start_color='ff0000',
                          end_color='ff0000',
                          fill_type='solid')
    sheet.cell(rowNum, columnNum).fill = redFill
    workBook.save(file)
